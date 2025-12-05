import azure.functions as func
import logging
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import tempfile
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
from datetime import datetime
import os


POSTGRES_URL_RMS = os.getenv("POSTGRES_URL_RMS")
POSTGRES_URL_DMS = os.getenv("POSTGRES_URL_DMS")
SLACK_TOKEN = os.getenv("SLACK_TOKEN")
SLACK_CHANNEL = os.getenv("SLACK_CHANNEL")

engine_RMS = create_engine(POSTGRES_URL_RMS)
engine_DMS = create_engine(POSTGRES_URL_DMS)

slack_client = WebClient(token=SLACK_TOKEN)


from .queries import TABLES



def run_sql(db_name, sql):

    if db_name.upper() == "RMS":
        engine = engine_RMS
    elif db_name.upper() == "DMS":
        engine = engine_DMS
    else:
        raise ValueError(f"Unknown DB: {db_name}")

    return pd.read_sql(sql, engine)



def build_excel_files():

    files = {}

    for db_name in ["RMS", "DMS"]:  

        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name, queries in TABLES.items():

            ws = wb.create_sheet(title=sheet_name)

            
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S IST")
            ws.append([f"Report generated on: {timestamp}"])
            ws.cell(row=1, column=1).font = Font(bold=True)
            ws.append([])

           
            for query_title, sql in queries.items():

                ws.append([f"• {query_title}"])
                ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

                try:
                    df = run_sql(db_name, sql)
                except Exception as e:
                    df = pd.DataFrame([{"ERROR": str(e)}])

                if df.empty:
                    ws.append(["No records found"])
                else:
                    for row in dataframe_to_rows(df, index=False, header=True):
                        ws.append(row)

                ws.append([])

        
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f"_{db_name}.xlsx")
        wb.save(tmp.name)
        files[db_name] = tmp.name

    return files["RMS"], files["DMS"]



def upload_both_to_slack(rms_path, dms_path):

   
    slack_client.files_upload_v2(
        channel=SLACK_CHANNEL,
        file=rms_path,
        filename="RMS_Report.xlsx",
        title="RMS Report"
    )

    
    slack_client.files_upload_v2(
        channel=SLACK_CHANNEL,
        file=dms_path,
        filename="DMS_Report.xlsx",
        title="DMS Report"
    )



async def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Combined RMS + DMS Report Triggered")

    try:
        rms_file, dms_file = build_excel_files()
        upload_both_to_slack(rms_file, dms_file)

        
        os.remove(rms_file)
        os.remove(dms_file)

        return func.HttpResponse("RMS & DMS Excel Reports Sent!", status_code=200)

    except Exception as e:
        logging.error(str(e))
        return func.HttpResponse(f"Error: {str(e)}", status_code=500)
